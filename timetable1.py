from openpyxl import load_workbook
from telegraphapi import Telegraph
import datetime
teleg = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,0]
teleg1 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,0]
moo = {}
gmu = {}
lin = {}
reks={}
par={}
days=["ПОНЕДЕЛЬНИК","ВТОРНИК","СРЕДА","ЧЕТВЕРГ","ПЯТНИЦА","СУББОТА","ВОСКРЕСЕНЬЕ"]
def start(a):
    file = a
    wb = load_workbook(file)

    days = []
    para =1
    par[1] = '1️⃣'
    par[2] = '2️⃣'
    par[3] = '3️⃣'
    par[4] = '4️⃣'
    par[5] = '5️⃣'
    mo,gm , ln,rek = [], [], [],[]
    dmo, dgm, dln, drek = [], [], [],[]
    k = 1
    d = 1
    sheets = [wb[' 1 '], wb[' 2 '], wb[' 3 '], wb[' 4 ']]

    for sheet in sheets:
        for j in range(2,15,2):
            for i in range(6,11):
                a=str(sheet.cell(row=i, column=j).value)
                b=str(sheet.cell(row=i, column=j+1).value)
                c=str(sheet.cell(row=i+8, column=j).value)
                d=str(sheet.cell(row=i+8, column=j+1).value)
                e=str(sheet.cell(row=i+16, column=j).value)
                f=str(sheet.cell(row=i+16, column=j+1).value)
                m=par[para]
                a=normalize(a,b,m)
                c=normalize(c,d,m)
                e=normalize(e,f,m)
                para+=1
                if a:
                    mo.append(a)
                if c:
                    gm.append(c)
                if e:
                    ln.append(e)
            para=1
            if mo:
                dmo.append(mo.copy())
            else:
                dmo.append("ВЫХОДНОЙ")
            if gm:
                dgm.append(gm.copy())
            else:
                dgm.append("ВЫХОДНОЙ")
            if ln:
                dln.append(ln.copy())
            else:
                dln.append("ВЫХОДНОЙ")
            #dche.append(chi.copy())
            mo.clear()
            gm.clear()
            ln.clear()
        moo[k] = dmo.copy()
        gmu[k] = dgm.copy()
        lin[k] = dln.copy()
        dmo.clear()
        dgm.clear()
        dln.clear()
        k+=1
def get_day(napr,kurs,d):
    #d=datetime.date.today().isoweekday()
    #print(d)
    a = ''
    if napr=="МО":
        a=moo[kurs][d-1]
    if napr=="ГМУ":
        a=gmu[kurs][d-1]
    if napr=="ЛИНГВ":
        a=lin[kurs][d-1]
    text=''
    if a=="ВЫХОДНОЙ":
        return  napr + '  -  '+ days[d-1] +"\n"+ a+ '\n'
    else:
        for i in range(0,len(a)):
            text= text +a[i]+ '\n'
        text = text.replace('[ПЗ]', ' -ПЗ- ')
        text = text.replace('[ЛК]', ' -ЛК- ')
        text = text.replace('[семинар]', ' -СЕМИНАР- ')
        text = text.replace('[//ЭКЗАМЕН]', ' -ЭКЗАМЕН- ')
        text = text.replace('[//ЗАЧЕТ]', ' -ЗАЧЁТ- ')
        text = text.replace('Фак:', '')
        text = text.replace('[//КОНСУЛЬТАЦИЯ]', ' -КОНСУЛЬТАЦИЯ- ')
        text = text.replace('-ПЗ-  *Стадион "Спартак"*', 'cтад. "Спартак"')
        # p = text.rindex('(')
        # p1 = text.rindex(')')
        # m = text[p + 1:p1 - 1]
        # print(m)
        # if len(m) > 11:
        #    k = text[:p1-1] + '\n    ' + text[p1:]
        #    print(k)
        #    text = k
        text = text.replace('(', '\n     ')
        text = text.replace(')', '')
        text = text.replace("-ауд", " каб")

        return napr + '  -  '+days[d-1] +"\n"+text
#get_day('ПМИИ',3)

def get_week(napr,kurs):

    k=0
    k=k+kurs
    if napr=="ГМУ":
       k=4+kurs
    elif napr=="ЛИНГВ":
        k=8+kurs
    if teleg[k]==0:
        a=''
        for i in range(0,7):
            a = a + "\n" + get_day(napr, kurs, i + 1)
        a = a.replace("*", "")
        a = a.replace("МО", "")
        a = a.replace("ГМУ", "")
        a = a.replace("ЛИНГВ", "")
        a = a.replace("- ", "")


        try:
            t = Telegraph()
            t.createAccount("PythonTelegraphAPI")
            title = napr + ' ' + str(kurs) + ' - Расписание'
            a += '\n<strong>Удачной учёбы! by \n <a href="t.me/MGURASP_Bot" target="_blank">t.me/MGURASP_Bot</a>'
            page = t.createPage(title, html_content=a)
            teleg[k]='http://telegra.ph/{}'.format(page['path'])
            return teleg[k]
        except:
            print("Error adding to telegra.ph")
            return 'В связи с блокировкой телеграмм произошла ошибка\n' + a
    else:
        return teleg[k]
def get_last_day(napr,kurs,dn):
    days=["ПОНЕДЕЛЬНИК","ВТОРНИК","СРЕДА","ЧЕТВЕРГ","ПЯТНИЦА","СУББОТА","ВОСКРЕСЕНЬЕ"]
    k=0
    k=k+kurs
    if napr=="ГМУ":
       k=4+kurs
    elif napr=="ЛИНГВ":
        k=8+kurs
    print(k)
    if teleg1[k]==0:
        a=''
        d = datetime.datetime.now()
        if d.hour<17:
            dnn=8
            ms=0
        else:
            dnn=7
            ms=1
        for i in range(dn,dnn):
             a= a +"\n" +get_day(napr,kurs,i+ms)
        a=a.replace("*","")
        a = a.replace("МО", "")
        a = a.replace("ГМУ", "")
        a = a.replace("ЛИНГВ", "")
        a = a.replace("- ", "")



        try:

            t = Telegraph()
            t.createAccount("PythonTelegraphAPI")
            title = napr + ' ' + str(kurs) + ' - Расписание'
            a += '\n<strong>Удачной учёбы! by \n <a href="t.me/MGURASP_Bot" target="_blank">t.me/MGURASP_Bot</a>'
            page = t.createPage(title, html_content=a)
            teleg1[k]='http://telegra.ph/{}'.format(page['path'])
            return teleg1[k]
        except:
            print("Error adding to telegra.ph")
            return 'В связи с блокировкой телеграмм произошла ошибка\n'+a
    else:
        return teleg1[k]

def normalize(a,b,para):
    if a!='None' :
        if b=='None' and a[0]!='"':
            return str(para)+' '+a
        elif b=="стд":
            return (str(para)+' '+a + ' *Стадион "Спартак"*')
        elif a[0]=='"':
            return a
        else:
            return (str(para)+' '+a+ ' в '+ '*'+b+' каб*')
    else:
        return 0


