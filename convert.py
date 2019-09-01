import xlrd
import imaplib
import email
import rarfile
import os
from openpyxl.workbook import Workbook
import smtplib

def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0,len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)
def get_email():
    username='afellay_maruf@mail.ru'
    password='afellay20sb'


    gmail=imaplib.IMAP4_SSL('imap.mail.ru', '993')
    gmail.login(username,password)
    #print(gmail.list())
    #print(gmail.select("INBOX"))
    stat,cnt=gmail.select('Schedule')

    stat, dta=gmail.uid('search', None, 'ALL')
    #    (cnt[0]),\
    #        ('UID BODY[TEXT])')
    #print(dta[0])
    inbox_list=dta[0].split()
    oldest=inbox_list[-1]
    res, data=gmail.uid('fetch', oldest, '(RFC822)')

    raw_m=data[0][1].decode("utf-8")
    email_mes=email.message_from_string(raw_m)
    a=email_mes
    #print(a)
    #b = get_first_text_block(email_mes)
    #print(b)
    mail = email.message_from_bytes(data[0][1])

    if mail.is_multipart():
        for part in mail.walk():
            content_type = part.get_content_type()
            filename = part.get_filename()
            if filename:
                # Нам плохого не надо, в письме может быть всякое барахло
                with open(part.get_filename(), 'wb') as new_file:
                    new_file.write(part.get_payload(decode=True))
    gmail.logout()
    z = rarfile.RarFile('Schedule.rar', 'r')
    b=z.namelist()
    #print(b)
    #c=z.getinfo(b[0])
    #print(c.filename)
    #z.open(c.filename)
    a=os.path.abspath('UnRAR.exe')
    print(a)
    aa=len(a)

    rarfile.UNRAR_TOOL = 'Unrar'
    print(b[1],b[0])
    z.extract(b[1])
    z.extract(b[0])

    cvt_xls_to_xlsx(b[0],str(b[0])+'x')
    cvt_xls_to_xlsx(b[1],str(b[1])+'x')
    f = open(str(b[0])+'x', 'rb')
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email.utils import formatdate
    from email import encoders
    text='проверка'
    server = smtplib.SMTP('smtp.mail.ru', 587)
    server.starttls()
    server.login(username, password)
    # Create a text/plain message
    msg = MIMEMultipart()
    msg['Subject'] = 'Greetings'
    msg['From'] = username
    msg['To'] = username
    msg.attach(MIMEText(text))
    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(b[1]+'x', "rb").read())
    encoders.encode_base64(part)
    file1=str(b[1])
    file2=str(b[0])
    pos1=file1.index('- ')
    pos2=file2.index('- ')
    part.add_header('Content-Disposition', 'attachment; filename="ENF '+str(b[1][pos1+1:(file1.index('н'))])+'.xlsx'+'"')
    msg.attach(part)

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(b[0] + 'x', "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="GF '+str(b[0][pos2+1:(file2.index('н'))])+'.xlsx'+'"')
    msg.attach(part)
    server.sendmail(username,username, msg.as_string())
    server.quit()
    return (str(b[1])+'x'), (str(b[0])+'x')
if __name__ == '__main__':
   a,b=get_email()
   print(a,b)
   
